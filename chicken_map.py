#!/usr/bin/python3

"""A 3D coordinate mapping program for chicken research"""

'''
Copyright (C) 2023  Logan Orians, in affiliation with Purdue University's
Dr. Marisa Erasmus and Gideon Ajibola.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program. If not, see https://www.gnu.org/licenses/.
'''

# Date: 07/13/23

# requirements: pip3 install -r requirements.txt

# Run Program
# Windows:      py chicken_map.py
# MacOS:        python3 chicken_map.py

# TODO:
# 1) add more error loggers
# 2) location-aware text, like how right-click menu flows up/down from cursor


__version__ = '2023.12.1'
__author__ = 'Logan Orians'


import argparse
import contextlib
import json
import logging
import os
import platform
#import re
import string
import time
import tkinter as tk
import types
from typing import Any, TypeVar

import cv2
import numpy as np
import openpyxl
import pytesseract # type: ignore

import options_gui

# Change working directory for .command executions
os.chdir(os.path.dirname(__file__))

# Custom Types for Type Checking (mypy)
TVideoCapture = TypeVar('TVideoCapture', bound=cv2.VideoCapture)


class FilePath:
    def __init__(self, directory: str) -> None:
        self.directory = self._make_proper_path(directory)
        self.filename = ''

    @staticmethod
    def _make_proper_path(directory: str) -> str:
        #lstrip() and rstrip()? needs testing
        if not directory.endswith('/') and not directory.endswith('\\'):
            directory += '/'
        if not os.path.exists(directory): os.makedirs(directory)

        return directory

    def _prevent_filename_overwrite(self) -> None:
        """Sets next available filename to avoid overwrite at same timestamp."""

        root, ext = os.path.splitext(self.filename) #get name and extension
        num = 0
        while os.path.exists(self.directory + self.filename):
            num += 1
            self.filename = f"{root}_{num}{ext}"

    def __str__(self) -> str:
        return f"{self.directory}{self.filename}"


class SpreadSheet(FilePath):
    def __init__(self, directory: str, fname: str, headers: list[str]) -> None:
        super().__init__(directory)
        self.filename = f"{fname}.xlsx"
        self._set_up_spreadsheet(headers)

    def _set_up_spreadsheet(self, headers: list[str]):
        """Sets up the output spreadsheet by adding bolded headers to columns.

        Args:
            headers: the bolded column headers to be written
        """

        with contextlib.closing(openpyxl.workbook.Workbook()) as wb:
            ws = wb.active
            ws.append(headers) # type: ignore[union-attr]
            for cell in ws['1:1']: # type: ignore[index]
                cell.font = openpyxl.styles.Font(bold=True) #make headers bold
            wb.save(str(self))

    def append_to_spreadsheet(self, data: list[str]):
        """Appends input data to spreadsheet.

        Args:
            data: the coord and timestamp to be appended
        """

        with contextlib.closing(openpyxl.load_workbook(str(self))) as wb:
            wb.active.append(data) # type: ignore[union-attr]
            wb.save(str(self))

    def delete_last_coordinate(self):
        """Deletes most recent coordinate from Excel sheet."""

        with contextlib.closing(openpyxl.load_workbook(str(self))) as wb:
            ws = wb.active
            last_row = ws.max_row
            ws.delete_rows(last_row)
            wb.save(str(self))


class AnnotationManager(FilePath):
    def __init__(self, directory: str) -> None:
        super().__init__(directory)
        self.anno_pos = (0, 0)
        self.anno_text = ''
        self.enter_time = 0.0
        self.filename = ''
        self.frame = None
        self.show_anno = False
        self.timestamp_time = ''
        self.typing = False
        self.write_anno = False

    def start_typing(self, x: int, y: int, timestamp_time: str) -> None:
        self.anno_pos = (x, y)
        self.anno_text = ''
        self.enter_time = 0.0
        self.filename = f"{timestamp_time.replace(':', '-')}.jpg"
        self._prevent_filename_overwrite()
        self.show_anno = True
        self.timestamp_time = timestamp_time
        self.typing = True
        self.write_anno = False


class ScreenCapture(FilePath):
    def __init__(self, directory: str) -> None:
        super().__init__(directory)
        self.captured= False
        self.capture_time = 0.0

    def save_frame(self, frame: Any, timestamp_time: str) -> None:
        self.captured = True
        self.capture_time = time.time()
        self.filename = f"{timestamp_time.replace(':', '-')}.jpg"
        #self._prevent_filename_overwrite()
        cv2.imwrite(str(self), frame)


class CoordinateManager():
    def __init__(self, three_d: bool | str) -> None:
        self.coord = () # type: tuple[int, ...]
        self.start_time = 0.0
        self.three_d = three_d
        if self.three_d == 'Floor':
            self.coord_3d = ()

    def set_coord(self, x, y):
        self.coord = (x, y)
        if self.three_d == 'Floor':
            self.coord_3d = self._get_3d_from_floor(x, y)

    @staticmethod
    def _get_3d_from_floor(x, y) -> tuple[int, int, int]:
        #camera view:
        # length: 10.54m
        # height: 2.57m (end), 2.38m (middle), 2.07m (close)
        #actual:
        # length: 12.1m
        # height: 2.2m
        # width minus nesting boxes: 3.04m
        # nesting boxes: 0.51m

        #real-world measurements, in meters
        x_meas = 3.04 #total floor width without nesting boxes
        y_meas = 10.54 #total floor length without nesting boxes
        nesting_boxes = 0.51

        #floor bounding box coords, in pixels
        # [1185, 200]  [1480, 185]  [2475, 1520]  [1030, 1520]
        #floor bounding box, in pixels
        floor_bb_x_min = 1030
        floor_bb_x_max = 2475
        floor_bb_y_min = 200
        floor_bb_y_max = 1520

        #nesting boxes bounding box coords, in pixels
        # [1030, 130]  [0, 1520]  [1030, 1520]  [1185, 200]
        nb_bb_x_min = 0
        nb_bb_x_max = 1185
        nb_bb_y_min = 130
        nb_bb_y_max = 1520

        #single roost bounding box, in pixels
        # [1650, 480]  [1820, 465]  [2470, 1050]  [2060, 1310]
        sr_bb_x_min = 1650
        sr_bb_x_max = 2470
        sr_bb_y_min = 465
        sr_bb_y_max = 1310

        #double roost bounding box, in pixels
        # [1360, 100]  [1480, 80]  [1780, 390]  [1625, 410]
        dr_bb_x_min = 1360
        dr_bb_x_max = 1780
        dr_bb_y_min = 80
        dr_bb_y_max = 410

        #polygon ROI - floor
        vertices_floor = np.array([[1185, 200], [1480, 185],
                                   [2475, 1520], [1030, 1520]], np.int32)
        mask_floor = np.zeros((1520, 2688), dtype=np.uint8)
        cv2.fillPoly(mask_floor, [vertices_floor], 255)

        #polygon ROI - nesting boxes
        vertices_nb = np.array([[1030, 130], [0, 1520],
                             [1030, 1520], [1185, 200]], np.int32)
        mask_nb = np.zeros((1520, 2688), dtype=np.uint8)
        cv2.fillPoly(mask_nb, [vertices_nb], 255)

        #polygon ROI - single roost
        vertices_sr = np.array([[1650, 480], [1820, 465],
                                [2470, 1050], [2060, 1310]], np.int32)
        mask_sr = np.zeros((1520, 2688), dtype=np.uint8)
        cv2.fillPoly(mask_sr, [vertices_sr], 255)

        #polygon ROI - double roost
        vertices_dr = np.array([[1360, 100], [1480, 80],
                                [1780, 390], [1625, 410]], np.int32)
        mask_dr = np.zeros((1520, 2688), dtype=np.uint8)
        cv2.fillPoly(mask_dr, [vertices_dr], 255)


        if mask_floor[y, x] == 255: #floor
            floor_bb_x = floor_bb_x_max - floor_bb_x_min
            floor_bb_y = floor_bb_y_max - floor_bb_y_min

            #scaling factor
            x_scale = x_meas / floor_bb_x
            y_scale = y_meas / floor_bb_y

            #normalization
            matrix = np.load('.3D_matrices/floor_matrix.npy')
            pixel_coord = np.array([x, y, 1])
            trans_pixel = matrix.dot(pixel_coord)
            trans_pixel /= trans_pixel[2]

            #scaling
            real_x = trans_pixel[0] * x_scale + nesting_boxes
            real_y = trans_pixel[1] * y_scale
            est_z = 0.2 #a chicken is a solid 40 cm tall, so halve that?

        elif mask_nb[y, x] == 255: #nesting boxes
            nb_bb_x = nb_bb_x_max - nb_bb_x_min
            nb_bb_y = nb_bb_y_max - nb_bb_y_min

            #scaling factor
            x_scale = nesting_boxes / nb_bb_x
            y_scale = y_meas / nb_bb_y

            #normalization
            matrix = np.load('.3D_matrices/nb_matrix.npy')
            pixel_coord = np.array([x, y, 1])
            trans_pixel = matrix.dot(pixel_coord)
            trans_pixel /= trans_pixel[2]

            #scaling
            real_x = trans_pixel[0] * x_scale
            real_y = trans_pixel[1] * y_scale
            est_z = 0.2 #a chicken is a solid 40 cm tall, so halve that?

        elif mask_sr[y, x] == 255: #single roost
            sr_bb_x = sr_bb_x_max - sr_bb_x_min
            sr_bb_y = sr_bb_y_max - sr_bb_y_min

            #scaling factor
            x_scale = 1 / sr_bb_x
            y_scale = 3.54 / sr_bb_y

            #normalization
            matrix = np.load('.3D_matrices/sr_matrix.npy')
            pixel_coord = np.array([x, y, 1])
            trans_pixel = matrix.dot(pixel_coord)
            trans_pixel /= trans_pixel[2]

            #scaling
            real_x = trans_pixel[0] * x_scale + nesting_boxes + 2
            real_y = trans_pixel[1] * y_scale + 7
            est_z = 0.4

        elif mask_dr[y, x] == 255: #double roost
            dr_bb_x = dr_bb_x_max - dr_bb_x_min
            dr_bb_y = dr_bb_y_max - dr_bb_y_min

            #scaling factor
            x_scale = 1 / dr_bb_x
            y_scale = 6.5 / dr_bb_y

            #normalization
            matrix = np.load('.3D_matrices/dr_matrix.npy')
            pixel_coord = np.array([x, y, 1])
            trans_pixel = matrix.dot(pixel_coord)
            trans_pixel /= trans_pixel[2]

            #scaling
            real_x = trans_pixel[0] * x_scale + nesting_boxes + 2
            real_y = trans_pixel[1] * y_scale
            est_z = 0.6

        else: #outside of my predefined boxes -> probably a wall or something
            real_x = -1
            real_y = -1
            est_z = -1

        return real_x, real_y, est_z


def mouse_input(
    event: int, x: int, y: int, flags: int,
    param: tuple[CoordinateManager, AnnotationManager, SpreadSheet])-> None:
    """Mouse input callback function for cv2.

    Args:
        event: mouse event (left- or right-click, etc.)
        x: x-coordinate of mouse cursor position
        y: y-coordinate of mouse cursor position
        flags: key event pressed with mouse input (shift, alt, etc.)
        param: parameters passed in to callback function by programmer, not cv2
    """

    del flags # Unused.
    coord, anno, sheet = param #unpack objects

    if not anno.typing: #if user isn't typing annotation
        if event == cv2.EVENT_LBUTTONDOWN: #left mouse click
            coord.start_time = time.time()
            coord.set_coord(x, y)
            timestamp_date, timestamp_time = get_timestamp(anno.frame)

            # Format data
            if coord.three_d == 'Floor':
                new_x, new_y, new_z = coord.coord_3d
                if new_x == -1:
                    data = [timestamp_date,timestamp_time, f"({x}, {y})", '( )']
                else:    
                    data = [timestamp_date, timestamp_time, f"({x}, {y})",
                            f"({new_x:.2f}, {new_y:.2f}, {new_z:.2f})"]
            else:
                data = [timestamp_date, timestamp_time, f"({x}, {y})"]
            sheet.append_to_spreadsheet(data)

            # Print timestamp and coordinates in case .xlsx gets corrupted
            print(timestamp_date)
            print(timestamp_time)
            print(f"{str(coord.coord)}\n")

        elif event == cv2.EVENT_RBUTTONDOWN: #right mouse click
            _, timestamp_time = get_timestamp(anno.frame)
            anno.start_typing(x, y, timestamp_time)


def get_timestamp(frame) -> tuple[str, str]:
    """Gets burnt-in timestamp via OCR (not video timestamp from OpenCV).

    Returns:
        timestamp_date: date from timestamp, DD/MM/YYYY
        timestamp_time: time from timestamp, HH:MM:SS
    """

    ts_area = frame[30:100, 26:634] #bounding box, [y:y+h, x:x+w]
    ts_gray = cv2.cvtColor(ts_area, cv2.COLOR_BGR2GRAY)
    #binary threshold for better recognition
    _, timestamp_thresh = cv2.threshold(ts_gray, 187, 255, cv2.THRESH_BINARY)
    #cv2.imshow('thresh', timestamp_thresh)
    #convert text in image to string
    timestamp = pytesseract.image_to_string(timestamp_thresh, config='--psm 7')
    #remove space, split after date
    timestamp_date, timestamp_time = timestamp.strip().split(' ')

    # regex in case it messes up. but it seems to be okay without it
    #rePattern = r'(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})' #regex group pattern
    #patternMatch = re.search(rePattern, timestamp)
    #timestamp_date = patternMatch.group(0)
    #timestamp_time = patternMatch.group(1)

    return timestamp_date, timestamp_time


def arg_parsing() -> argparse.Namespace:
    """Parses input arguments.

    Returns:
        parser.parse_args(): object containing command line args
    """

    parser = argparse.ArgumentParser(description=('A program that displays and '
        'saves coordinates for tracking chicken behavior.'))
    parser.add_argument('-o', '--options', action='store_true',
        help='Opens the GUI for setting program options.')
    parser.add_argument('--version', action='version',
        version=f"%(prog)s {__version__}", help=argparse.SUPPRESS)

    return parser.parse_args()


def get_args_from_file(filename: str) -> dict[str, Any]:
    """Gets program options from file.

    Args:
        filename: name of file that contains options

    Returns:
        args: arguments from file
    """

    try:
        with open(filename, 'r') as f:
            args = json.load(f)
        return args
    except OSError as e:
        print(f'Error accessing file; please contact author: {e}')
        raise
    except json.JSONDecodeError as e:
        print(f'Error decoding JSON, please contact author: {e}')
        raise


# TODO: get more info?
def get_system_info() -> str:
    """Gets system info, formatted into lines, for the error log file.

    Returns:
        info: multi-line string containing system info
    """

    uname = platform.uname()
    info = (
        f"\nOS: {uname.system}\n"
        f"Release: {uname.release}\n"
        f"Version: {uname.version}\n"
        f"Processor: {uname.processor}\n"
        f"Python: {platform.python_version()}")

    return info


def set_up_logger() -> logging.Logger:
    """Sets up error logger with custom terminator to separate error entries.

    Returns:
        logger (logging.Logger): an instance of a logger object
    """

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler('error_log.txt')
    file_handler.terminator = f"\n\n{'=' * 80}\n\n"
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    return logger


def get_window_and_video_dims(cap: TVideoCapture) -> tuple[int, int, int, int]:
    """Gets screen resolution and returns suitable window dimensions.

    Args:
        cap (cv2.VideoCapture obj): video captured by OpenCV

    Returns:
        window_width: suitable window width, in pixels
        window_height: suitable window height, in pixels
        video_width: video resolution width, in pixels
        video_height: video resolution height, in pixels
    """

    # Get screen resolution
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.destroy()

    video_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    video_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    aspect_ratio = video_width / video_height
    window_width = int(min(screen_width - 150, video_width)) #taskbar, etc.
    window_height = int(window_width / aspect_ratio)

    return window_width, window_height, video_width, video_height


def key_ascii(key: str) -> int:
    """Get ASCII value of key (options_gui only allows printable chars + Esc).

    Args:
        key: key to get ASCII value of

    Returns:
        ASCII value of key
    """

    if key.lower() == 'esc': return 27
    return ord(key)


def main():
    
    args = arg_parsing()
    if args.options:
        options_gui.main() #run GUI

    logger = set_up_logger() #set up bad error logger

    ascii_allowlist = string.printable[:-5] #OpenCV can only print up to <space>
    system_date_time = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())

    #point pytesseract to tesseract executable
    if platform.system() == 'Windows':
        pytesseract.pytesseract.tesseract_cmd = R'C:\Program Files\Tesseract-OCR\tesseract.exe'
    
    options_file = '.options.json'
    prog_options = types.SimpleNamespace(**get_args_from_file(options_file))

    # Set up arguments for program use
    infile_path = prog_options.video_path.strip() #strip whitespace for MacOS
    exit_key = key_ascii(prog_options.exit_key)
    clear_key = key_ascii(prog_options.clear_key)
    pause_key = key_ascii(prog_options.pause_key)
    screencap_key = key_ascii(prog_options.screencap_key)
    duration = prog_options.duration #duration on screen, in seconds
    font = types.SimpleNamespace(font=prog_options.font,
                                 scale=prog_options.font_scale,
                                 color=tuple(reversed(prog_options.font_color)),
                                 thickness=prog_options.font_thickness)

    # Instantiate classes
    coord = CoordinateManager(prog_options.three_d)
    anno = AnnotationManager(f"{prog_options.anno_dir}/{system_date_time}")
    screencap = ScreenCapture(
        f"{prog_options.screencaps_dir}/{system_date_time}")
    if coord.three_d == 'Floor':
        headers = ['Date', 'Time', 'Coordinates', '3D Coordinates']
    else:
        headers = ['Date', 'Time', 'Coordinates']
    sheet = SpreadSheet(prog_options.out_dir, system_date_time, headers)

    # Determine delay to play video at normal speed
    cap = cv2.VideoCapture(infile_path) #create Video Capture object
    fps = cap.get(cv2.CAP_PROP_FPS) #get fps of cap input
    if fps == 0:
        fps = 25 #set default if determination fails
    delay = int(1000 / fps) #calculate delay from fps, in ms
    '''
    Note about delay: It's about 20% slower than real time...even though
    this 100% should work. Might have a possible fix in the works.
    '''

    # Set up video window
    w_width, w_height, v_width, v_height = get_window_and_video_dims(cap)
    window_name = 'Video'
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL) #named window to display cap
    cv2.resizeWindow(window_name, width=w_width, height=w_height)

    paused = False
    callback_params = coord, anno, sheet
    cv2.setMouseCallback(window_name, mouse_input, param=callback_params)


    try:
        while cap.isOpened():
            #frame_start_time = time.time()

            if not paused:
                ret, anno.frame = cap.read() #get cap frame-by-frame
                if not ret: break

            #timing while True here, everything else gets indented?
            #might need to rework pause
                
                # Get LSByte of keypress for cross-platform compatibility
                key_press = cv2.waitKey(delay) & 0xFF
                if not anno.typing:
                    if key_press == exit_key: #quit program
                        break
                    if anno.write_anno:
                        anno.write_anno = False
                        cv2.imwrite(str(anno), frame_copy)
                    if time.time() - anno.enter_time > duration:
                        anno.show_anno = False
                        anno.anno_text = ''

                if not screencap.captured:
                    if key_press == screencap_key:
                        screencap.save_frame(anno.frame,
                                             get_timestamp(anno.frame)[1])
                if screencap.captured:
                    cv2.putText(anno.frame, 'Screencap saved!', (500, 500),
                                font.font, font.scale,
                                font.color, font.thickness)
                    if time.time() - screencap.capture_time > 1.15:
                        screencap.captured = False

            # Prevent coords from popping back up after annotation is entered
            if anno.typing: coord.coord = ()

            # Only allow deletion of [date, time, coord] when on screen
            if coord.coord:
                frame_copy = anno.frame.copy() #clearing looks better
                if key_press == clear_key:
                    coord.coord = ()
                    sheet.delete_last_coordinate()
                elif (time.time() - coord.start_time < duration): #coord timeout
                    if paused:
                        cv2.putText(frame_copy, str(coord.coord),
                                    coord.coord, font.font, font.scale,
                                    font.color, font.thickness)
                    else:
                        cv2.putText(anno.frame, str(coord.coord),
                                    coord.coord, font.font, font.scale,
                                    font.color, font.thickness)

            # This while loop ensures that the video is paused while annotating
            while anno.typing:
                frame_copy = anno.frame.copy() #makes backspace work when typing

                if anno.show_anno:
                    cv2.putText(frame_copy, anno.anno_text, anno.anno_pos,
                                font.font, font.scale, font.color,
                                font.thickness)
                    cv2.imshow(window_name, frame_copy)

                key_press = cv2.waitKey(0) & 0xFF #LSByte for cross-plat compat
                if key_press != 255:
                    if key_press == 13: #Enter
                        anno.typing = False
                        anno.write_anno = True
                        anno.enter_time = time.time()
                    elif key_press == 27: #Esc
                        anno.typing = False
                        anno.show_anno = False
                        anno.anno_text = ''
                    elif key_press == 8 or key_press == 127: #Backspace/Del
                        if anno.anno_text: #DON'T COMBINE INTO ^ELIF
                            anno.anno_text = anno.anno_text[:-1]
                    elif chr(key_press) in ascii_allowlist: #printable ascii
                        anno.anno_text += chr(key_press)
                        anno.show_anno = True

            if anno.show_anno:
                if paused:
                    cv2.putText(frame_copy, anno.anno_text, anno.anno_pos,
                                font.font,font.scale, font.color,
                                font.thickness)
                else:
                    cv2.putText(anno.frame, anno.anno_text, anno.anno_pos,
                                font.font, font.scale, font.color,
                                font.thickness)

            if not(paused and (coord.coord or anno.show_anno)):
                cv2.imshow(window_name, anno.frame) #show video frame


    except Exception as e:
        logger.error('\nError: %s\n', e, exc_info=True)
        print(('\n***AN ERROR OCCURRED! '
               'PLEASE FOLLOW THE SUPPORT INSTRUCTIONS IN THE README.***'))

    finally:
        # destroy cv2 windows if initialized
        if 'cap' in locals() or 'cap' in globals():
            cap.release() #release video capture object
            cv2.destroyAllWindows() #close all OpenCV windows


if __name__ == "__main__":
    main()
