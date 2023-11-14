#!/usr/bin/python3

#TODO: follow google style guide docstrings (incorporate Usage)
"""A coordinate mapping program made initially for chicken research"""

'''
Copyright 2023, Logan Orians in affiliation with Purdue University:
    Dr. Marisa Erasmus and Gideon Ajibola.

Approved for private use by students and employees of Purdue University only.
No implied support or warranty.
'''

# Date: 07/13/23

# requirements: pip3 install -r requirements.txt

# Run Program
# Windows:      py chickenMap.py
# MacOS:        python3 chickenMap.py

# TODO:
# 1) .bat/.sh/.command files for double-click running (no cmd/terminal)
# 2) shift-left click to add a note to the spreadsheet - use flags in mouse_cb
# 3) add more error loggers
# 4) location-aware text drawing; how right-click menu flows up/down from cursor
# 5) remove parentheses where unncessary


__version__ = '2023.11.1'
__author__ = 'Logan Orians'


import argparse
import contextlib
import json
import logging
import os
import platform
#import re
import string
import time
import tkinter as tk
import types
from typing import Any, TypeVar

import cv2
import openpyxl
import pytesseract # type: ignore

import options_gui


TVideoCapture = TypeVar('TVideoCapture', bound=cv2.VideoCapture)


class FilePath:
    def __init__(self, directory: str) -> None:
        self.directory = self._make_proper_path(directory)
        self.filename = ''

    @staticmethod
    def _make_proper_path(directory: str) -> str:
        #lstrip() and rstrip()? needs testing
        if not directory.endswith('/') and not directory.endswith('\\'):
            directory += '/'
        if not os.path.exists(directory): os.makedirs(directory)

        return directory

    def __str__(self) -> str:
        return f"{self.directory}{self.filename}"


class SpreadSheet(FilePath):
    def __init__(self, directory: str, fname: str, headers: list[str]) -> None:
        super().__init__(directory)
        self.filename = f"{fname}.xlsx"
        self._set_up_spreadsheet(headers)

    def _set_up_spreadsheet(self, headers: list[str]):
        """Sets up the output spreadsheet by adding bolded headers to columns.

        Args:
            headers: the bolded column headers to be written
        """

        with contextlib.closing(openpyxl.workbook.Workbook()) as wb:
            ws = wb.active
            ws.append(headers) # type: ignore[union-attr]
            for cell in ws['1:1']: # type: ignore[index]
                cell.font = openpyxl.styles.Font(bold=True) #make headers bold
            wb.save(str(self))

    def append_to_spreadsheet(self, data: list[str]):
        """Appends input data to spreadsheet.

        Args:
            data: the coord and timestamp to be appended
        """

        with contextlib.closing(openpyxl.load_workbook(str(self))) as wb:
            wb.active.append(data) # type: ignore[union-attr]
            wb.save(str(self))

    def delete_last_coordinate(self):
        """Deletes most recent coordinate from Excel sheet."""

        with contextlib.closing(openpyxl.load_workbook(str(self))) as wb:
            ws = wb.active
            last_row = ws.max_row
            ws.delete_rows(last_row)
            wb.save(str(self))


class AnnotationManager(FilePath):
    def __init__(self, directory: str) -> None:
        super().__init__(directory)
        self.anno_pos = (0, 0)
        self.anno_text = ''
        self.enter_time = 0.0
        self.filename = ''
        self.frame = None
        self.show_anno = False
        self.timestamp_time = ''
        self.typing = False
        self.write_anno = False

    def start_typing(self, x: int, y: int, timestamp_time: str) -> None:
        self.anno_pos = (x, y)
        self.anno_text = ''
        self.enter_time = 0.0
        self.filename = f"{timestamp_time.replace(':', '-')}.jpg"
        self._prevent_filename_overwrite()
        self.show_anno = True
        self.timestamp_time = timestamp_time
        self.typing = True
        self.write_anno = False

    def _prevent_filename_overwrite(self) -> None:
        """Sets next available filename to avoid overwrite at same timestamp."""

        root, ext = os.path.splitext(self.filename) #get name and extension
        num = 0
        while os.path.exists(self.directory + self.filename):
            num += 1
            self.filename = f"{root}_{num}{ext}"


class CoordinateManager():
    def __init__(self, three_d: bool | str) -> None:
        self._coord = () # type: tuple[int, ...]
        self.start_time = 0.0
        self._three_d = three_d

    @property
    def coord(self) -> tuple[int, ...]:
        return self._coord
    
    @coord.setter
    def coord(self, x_and_y: tuple[int, ...]) -> None:
        if self._three_d:
            self._coord = self._get_3d_from_2d(*x_and_y)
        else:
            self._coord = x_and_y

    @staticmethod
    def _get_3d_from_2d(x, y) -> tuple[int, int]:
        return (x, y)


def mouse_input(
    event: int, x: int, y: int, flags: int,
    param: tuple[CoordinateManager, AnnotationManager, SpreadSheet])-> None:
    """Mouse input callback function for cv2.

    Args:
        event: mouse event (left- or right-click, etc.)
        x: x-coordinate of mouse cursor position
        y: y-coordinate of mouse cursor position
        flags: key event pressed with mouse input (shift, alt, etc.)
        param: parameters passed in to callback function by programmer, not cv2
    """

    del flags # Unused.
    coord, anno, sheet = param #unpack objects

    if not anno.typing: #if user isn't typing annotation
        if event == cv2.EVENT_LBUTTONDOWN: #left mouse click
            coord.start_time = time.time()
            coord.coord = (x, y)
            timestamp_date, timestamp_time = get_timestamp(anno.frame)

            data = [timestamp_date, timestamp_time, f"({x}, {y})"] #format data
            sheet.append_to_spreadsheet(data)

            # Print timestamp and coordinates in case .xlsx gets corrupted
            print(timestamp_date)
            print(timestamp_time)
            print(f"{str(coord.coord)}\n")

        elif event == cv2.EVENT_RBUTTONDOWN: #right mouse click
            _, timestamp_time = get_timestamp(anno.frame)
            anno.start_typing(x, y, timestamp_time)


def get_timestamp(frame) -> tuple[str, str]:
    """Gets burnt-in timestamp via OCR (not video timestamp from OpenCV).

    Returns:
        timestamp_date: date from timestamp, DD/MM/YYYY
        timestamp_time: time from timestamp, HH:MM:SS
    """

    ts_area = frame[30:100, 26:634] #bounding box, [y:y+h, x:x+w]
    ts_gray = cv2.cvtColor(ts_area, cv2.COLOR_BGR2GRAY)
    #binary threshold for better recognition
    _, timestamp_thresh = cv2.threshold(ts_gray, 187, 255, cv2.THRESH_BINARY)
    #cv2.imshow('thresh', timestamp_thresh)
    #convert text in image to string
    timestamp = pytesseract.image_to_string(timestamp_thresh, config='--psm 7')
    #remove space, split after date
    timestamp_date, timestamp_time = timestamp.strip().split(' ')

    # regex in case it messes up. but it seems to be okay without it
    #rePattern = r'(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2}:\d{2})' #regex group pattern
    #patternMatch = re.search(rePattern, timestamp)
    #timestamp_date = patternMatch.group(0)
    #timestamp_time = patternMatch.group(1)

    return timestamp_date, timestamp_time


def arg_parsing() -> argparse.Namespace:
    """Parses input arguments.

    Returns:
        parser.parse_args(): object containing command line args
    """

    parser = argparse.ArgumentParser(description=('A program that displays and '
        'saves coordinates for tracking chicken behavior.'))
    parser.add_argument('-o', '--options', action='store_true',
        help='Opens the GUI for setting program options.')
    parser.add_argument('--version', action='version',
        version=f"%(prog)s {__version__}", help=argparse.SUPPRESS)

    return parser.parse_args()


def get_args_from_file(filename: str) -> dict[str, Any]:
    """Gets program options from file.

    Args:
        filename: name of file that contains options

    Returns:
        args: arguments from file
    """

    try:
        with open(filename, 'r') as f:
            args = json.load(f)
        return args
    except OSError as e:
        print(f'Error accessing file; please contact author: {e}')
        raise
    except json.JSONDecodeError as e:
        print(f'Error decoding JSON, please contact author: {e}')
        raise


# TODO: get more info?
def get_system_info() -> str:
    """Gets system info, formatted into lines, for the error log file.

    Returns:
        info: multi-line string containing system info
    """

    uname = platform.uname()
    info = (
        f"\nOS: {uname.system}\n"
        f"Release: {uname.release}\n"
        f"Version: {uname.version}\n"
        f"Processor: {uname.processor}\n"
        f"Python: {platform.python_version()}")

    return info


def set_up_logger() -> logging.Logger:
    """Sets up error logger with custom terminator to separate error entries.

    Returns:
        logger (logging.Logger): an instance of a logger object
    """

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler('error_log.txt')
    file_handler.terminator = f"\n\n{'=' * 80}\n\n"
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    return logger


def get_window_and_video_dims(cap: TVideoCapture) -> tuple[int, int, int, int]:
    """Gets screen resolution and returns suitable window dimensions.

    Args:
        cap (cv2.VideoCapture obj): video captured by OpenCV

    Returns:
        window_width: suitable window width, in pixels
        window_height: suitable window height, in pixels
        video_width: video resolution width, in pixels
        video_height: video resolution height, in pixels
    """

    # Get screen resolution
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.destroy()

    video_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    video_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    aspect_ratio = video_width / video_height
    window_width = int(min(screen_width - 150, video_width)) #taskbar, etc.
    window_height = int(window_width / aspect_ratio)

    return window_width, window_height, video_width, video_height


def main():
    
    args = arg_parsing()
    if args.options:
        options_gui.main() #run GUI

    logger = set_up_logger() #set up bad error logger

    ascii_allowlist = string.printable[:-5] #OpenCV can only print up to <space>
    system_date_time = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())

    #point pytesseract to tesseract executable
    if platform.system() == 'Windows':
        pytesseract.pytesseract.tesseract_cmd = R'C:\Program Files\Tesseract-OCR\tesseract.exe'
    
    options_file = '.options.json'
    prog_options = types.SimpleNamespace(**get_args_from_file(options_file))

    # Set up arguments for program use
    infile_path = prog_options.video_path.strip() #strip whitespace for MacOS
    if prog_options.exit_key.lower() == 'esc':
        exit_key = 27
    else:
        exit_key = ord(prog_options.exit_key)
    clear_key = ord(prog_options.clear_key)
    pause_key = ord(prog_options.pause_key)
    duration = prog_options.duration #duration on screen, in seconds
    font = types.SimpleNamespace(font=prog_options.font,
                                 scale=prog_options.font_scale,
                                 color=tuple(reversed(prog_options.font_color)),
                                 thickness=prog_options.font_thickness)

    # Instantiate classes
    #coord = types.SimpleNamespace(xy=(), start_time=0)
    coord = CoordinateManager(prog_options.three_d)
    anno = AnnotationManager(f"{prog_options.anno_dir}/{system_date_time}")
    headers = ['Date', 'Time', 'Coordinates']
    sheet = SpreadSheet(prog_options.out_dir, system_date_time, headers)

    # Determine delay to play video at normal speed
    cap = cv2.VideoCapture(infile_path) #create Video Capture object
    fps = cap.get(cv2.CAP_PROP_FPS) #get fps of cap input
    if fps == 0:
        fps = 25 #set default if determination fails
    delay = int(1000 / fps) #calculate delay from fps, in ms

    # Set up video window
    w_width, w_height, v_width, v_height = get_window_and_video_dims(cap)
    window_name = 'Video'
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL) #named window to display cap
    cv2.resizeWindow(window_name, width=w_width, height=w_height)

    paused = False
    callback_params = coord, anno, sheet
    cv2.setMouseCallback(window_name, mouse_input, param=callback_params)


    try:
        while cap.isOpened():
            if not paused:
                ret, anno.frame = cap.read() #get cap frame-by-frame
                if not ret: break
                
                # Get LSByte of keypress for cross-platform compatibility
                key_press = cv2.waitKey(delay) & 0xFF
                if not anno.typing:
                    if key_press == exit_key: #quit program
                        break
                    if anno.write_anno:
                        anno.write_anno = False
                        cv2.imwrite(str(anno), frame_copy)
                    if time.time() - anno.enter_time > duration:
                        anno.show_anno = False
                        anno.anno_text = ''

            # Prevent coords from popping back up after annotation is entered
            if anno.typing:
                coord.coord = ()

            # Only allow deletion of [date, time, coord] when on screen
            if coord.coord:
                frame_copy = anno.frame.copy() #clearing looks better
                if key_press == clear_key:
                    coord.coord = ()
                    sheet.delete_last_coordinate()
                elif (time.time() - coord.start_time < duration): #coord timeout
                    if paused:
                        cv2.putText(frame_copy, str(coord.coord),
                                    coord.coord[:2], font.font, font.scale,
                                    font.color, font.thickness)
                    else:
                        cv2.putText(anno.frame, str(coord.coord),
                                    coord.coord[:2], font.font, font.scale,
                                    font.color, font.thickness)

            # This while loop ensures that the video is paused while annotating
            while anno.typing:
                frame_copy = anno.frame.copy() #makes backspace work when typing

                if anno.show_anno:
                    cv2.putText(frame_copy, anno.anno_text, anno.anno_pos,
                                font.font, font.scale, font.color,
                                font.thickness)
                    cv2.imshow(window_name, frame_copy)

                key_press = cv2.waitKey(0) & 0xFF #LSByte for cross-plat compat
                if key_press != 255:
                    if key_press == 13: #Enter
                        anno.typing = False
                        anno.write_anno = True
                        anno.enter_time = time.time()
                    elif key_press == 27: #Esc
                        anno.typing = False
                        anno.show_anno = False
                        anno.anno_text = ''
                    elif key_press == 8: #Backspace
                        if anno.anno_text: #DON'T COMBINE INTO ^ELIF
                            anno.anno_text = anno.anno_text[:-1]
                    elif chr(key_press) in ascii_allowlist: #printable ascii
                        anno.anno_text += chr(key_press)
                        anno.show_anno = True

            if anno.show_anno:
                if paused:
                    cv2.putText(frame_copy, anno.anno_text, anno.anno_pos,
                                font.font,font.scale, font.color,
                                font.thickness)
                else:
                    cv2.putText(anno.frame, anno.anno_text, anno.anno_pos,
                                font.font, font.scale, font.color,
                                font.thickness)

            if not(paused and (coord.coord or anno.show_anno)):
                cv2.imshow(window_name, anno.frame) #show video frame

    except Exception as e:
        logger.error('\nError: %s\n', e, exc_info=True)
        print(('\n***AN ERROR OCCURRED! '
               'PLEASE FOLLOW THE SUPPORT INSTRUCTIONS IN THE README.***'))

    finally:
        # destroy cv2 windows if initialized
        if 'cap' in locals() or 'cap' in globals():
            cap.release() #release video capture object
            cv2.destroyAllWindows() #close all OpenCV windows


if __name__ == "__main__":
    main()
